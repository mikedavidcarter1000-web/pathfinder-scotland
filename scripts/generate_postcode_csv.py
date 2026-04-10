#!/usr/bin/env python3
"""
Generate CSV file for SIMD postcodes that can be imported via Supabase Dashboard.
The CSV format matches the simd_postcodes table structure.
"""

import openpyxl
import os
import csv

def normalize_postcode(postcode):
    """Normalize postcode by removing spaces and converting to uppercase."""
    if postcode is None:
        return None
    return postcode.upper().replace(' ', '')

def main():
    excel_file = os.path.join(os.path.dirname(__file__), '..', 'simd_postcodes.xlsx')
    output_file = os.path.join(os.path.dirname(__file__), '..', 'data', 'simd_postcodes.csv')

    # Create data directory if needed
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"Loading Excel file: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['All postcodes']

    total_rows = ws.max_row - 1
    print(f"Total postcodes to process: {total_rows}")
    print(f"Writing to: {output_file}")

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        # Header matching simd_postcodes table
        writer.writerow(['postcode', 'simd_decile', 'datazone', 'council_area'])

        processed = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            postcode = row[0]
            datazone = row[1]
            simd_decile = row[4]  # SIMD2020_Decile column

            # Validate
            if not postcode or simd_decile is None:
                skipped += 1
                continue

            try:
                decile = int(simd_decile)
                if decile < 1 or decile > 10:
                    skipped += 1
                    continue
            except (ValueError, TypeError):
                skipped += 1
                continue

            normalized = normalize_postcode(postcode)
            writer.writerow([normalized, decile, datazone or '', ''])

            processed += 1

            if processed % 50000 == 0:
                print(f"Processed: {processed}/{total_rows} ({100*processed/total_rows:.1f}%)")

    print(f"\nDone! Processed: {processed}, Skipped: {skipped}")
    print(f"CSV file: {output_file}")
    print(f"\nTo import via Supabase Dashboard:")
    print("1. Go to Table Editor > simd_postcodes")
    print("2. Click 'Insert' > 'Import data from CSV'")
    print("3. Select the generated CSV file")
    print("4. Enable 'Upsert' mode to handle duplicates")

    wb.close()

if __name__ == "__main__":
    main()
