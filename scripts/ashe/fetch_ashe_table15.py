#!/usr/bin/env python3
"""
Fetch and parse ONS ASHE salary data by 4-digit SOC code.

Downloads two datasets:
  - ASHE Table 15 (region × occupation) → Scotland figures
  - ASHE Table 14 (UK-level occupation) → UK fallback figures

Produces two JSON files in scripts/ashe/:
  - ashe-2025-scotland.json
  - ashe-2025-uk.json

Structure:
  { "1111": { "description": "Chief executives ...", "p25": 68000, "p50": 92000, "p75": 140000 }, ... }
  Suppressed values ("..", "x", etc.) → null.

Source: ASHE 2025 provisional (released Oct 2025).
  Table 15 has no UK aggregate rows — only English regions, Scotland, Wales.
  Table 14 provides UK-level data with better percentile coverage.

Usage:
    python scripts/ashe/fetch_ashe_table15.py
"""

import json
import os
import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path
from urllib.request import urlopen, Request

import openpyxl

# ---------------------------------------------------------------------------
# URLs
# ---------------------------------------------------------------------------

TABLE15_URL = (
    "https://www.ons.gov.uk/file?uri=/employmentandlabourmarket/peopleinwork/"
    "earningsandworkinghours/datasets/regionbyoccupation4digitsoc2010ashetable15/"
    "2025provisional/ashetable152025provisional.zip"
)

TABLE14_URL = (
    "https://www.ons.gov.uk/file?uri=/employmentandlabourmarket/peopleinwork/"
    "earningsandworkinghours/datasets/occupation4digitsoc2010ashetable14/"
    "2025provisional/ashetable142025provisional.zip"
)

SCRIPT_DIR = Path(__file__).resolve().parent
SUPPRESSED = {"..", "x", ":", "-", "#", "~", "*", "c", "..c", ""}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def download_zip(url: str) -> BytesIO:
    print(f"Downloading: {url[:90]}...")
    req = Request(url, headers={"User-Agent": "PathfinderScotland/1.0 (education)"})
    with urlopen(req, timeout=180) as resp:
        data = resp.read()
    print(f"  {len(data) / 1024 / 1024:.1f} MB")
    return BytesIO(data)


def find_file_in_zip(zf: zipfile.ZipFile, *, soc_digits: str, table_num: str) -> str | None:
    """
    Find the annual-pay-gross Excel file for a given SOC digit level.
    e.g. soc_digits="4", table_num="15" → Table 15 (4).7a Annual pay
    """
    for name in zf.namelist():
        lower = name.lower()
        if (
            f"({soc_digits})" in lower
            and ".7a" in lower
            and "annual" in lower
            and lower.endswith(".xlsx")
            and "cv" not in lower
        ):
            return name
    # Fallback: any annual pay file with the right SOC level
    for name in zf.namelist():
        lower = name.lower()
        if f"({soc_digits})" in lower and "annual" in lower and lower.endswith(".xlsx") and "cv" not in lower:
            return name
    return None


def parse_salary(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return None
        return round(value)
    s = str(value).strip().lower()
    if s in SUPPRESSED:
        return None
    s = s.replace(",", "").replace("£", "").strip()
    if s in SUPPRESSED:
        return None
    try:
        return round(float(s))
    except ValueError:
        return None


def is_4digit_soc(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if "." in s:
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    return s if re.match(r"^\d{4}$", s) else None


def detect_header(ws) -> int:
    """Find the header row containing 'Description' and 'Code'."""
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=5, values_only=False):
        labels = [str(c.value).strip().lower() if c.value else "" for c in row]
        if "description" in labels and "code" in labels:
            return row[0].row
    return -1


def find_percentile_cols(ws, header_row: int) -> dict:
    """
    Read the header row and return column indices (0-based tuple index) for key fields.

    Known ASHE annual pay layout (confirmed from 2025 data):
      idx 0: Description     idx 1: Code       idx 2: Number of jobs
      idx 3: Median          idx 4: change     idx 5: Mean     idx 6: change
      idx 7: p10  idx 8: p20  idx 9: p25  idx 10: p30  idx 11: p40
      idx 12: p60  idx 13: p70  idx 14: p75  idx 15: p80  idx 16: p90
    """
    cols = {"desc": 0, "code": 1, "median": 3}  # defaults from known layout

    # Verify by reading the actual header
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, max_col=25, values_only=True):
        for idx, val in enumerate(row):
            if val is None:
                continue
            label = str(val).strip().lower()
            if label == "description":
                cols["desc"] = idx
            elif label == "code":
                cols["code"] = idx
            elif label == "median":
                cols["median"] = idx
            else:
                try:
                    pval = int(label)
                    if pval in (10, 20, 25, 30, 40, 60, 70, 75, 80, 90):
                        cols[f"p{pval}"] = idx
                except ValueError:
                    pass
        break

    return cols


# ---------------------------------------------------------------------------
# Table 15 parser (Scotland from regional data)
# ---------------------------------------------------------------------------

def parse_table15_scotland(zf: zipfile.ZipFile) -> dict:
    """Extract Scotland salary data from ASHE Table 15 (4-digit SOC)."""
    target = find_file_in_zip(zf, soc_digits="4", table_num="15")
    if not target:
        print("ERROR: Could not find Table 15 (4) annual pay file.")
        return {}

    print(f"Table 15 target: {os.path.basename(target)}")

    # Save raw Excel locally (gitignored)
    raw_path = SCRIPT_DIR / os.path.basename(target)
    with open(raw_path, "wb") as f:
        f.write(zf.read(target))

    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws = wb["All"]
    header_row = detect_header(ws)
    if header_row < 0:
        print("  ERROR: Header row not found in Table 15.")
        wb.close()
        return {}

    cols = find_percentile_cols(ws, header_row)
    print(f"  Header row {header_row}, columns: {cols}")

    scotland = {}
    c_desc = cols["desc"]
    c_code = cols["code"]
    c_median = cols["median"]
    c_p25 = cols.get("p25")
    c_p75 = cols.get("p75")

    for row in ws.iter_rows(min_row=header_row + 1, max_col=25, values_only=True):
        desc_raw = row[c_desc] if c_desc is not None and c_desc < len(row) else None
        if not desc_raw:
            continue

        desc_str = str(desc_raw).strip()
        if not desc_str.lower().startswith("scotland"):
            continue

        code_val = row[c_code] if c_code is not None and c_code < len(row) else None
        soc = is_4digit_soc(code_val)
        if not soc:
            continue

        # Extract occupation title (after "Scotland, ")
        parts = desc_str.split(",", 1)
        occ_title = parts[1].strip() if len(parts) == 2 else desc_str

        p25 = parse_salary(row[c_p25]) if c_p25 is not None and c_p25 < len(row) else None
        p50 = parse_salary(row[c_median]) if c_median < len(row) else None
        p75 = parse_salary(row[c_p75]) if c_p75 is not None and c_p75 < len(row) else None

        scotland[soc] = {"description": occ_title, "p25": p25, "p50": p50, "p75": p75}

    wb.close()
    return scotland


# ---------------------------------------------------------------------------
# Table 14 parser (UK-level data)
# ---------------------------------------------------------------------------

def parse_table14_uk(zf: zipfile.ZipFile) -> dict:
    """Extract UK-level salary data from ASHE Table 14 (4-digit SOC)."""
    target = find_file_in_zip(zf, soc_digits="4", table_num="14")
    if not target:
        # Table 14 might not have the (4) subfolder convention
        # Try finding any annual pay gross file
        for name in zf.namelist():
            lower = name.lower()
            if ".7a" in lower and "annual" in lower and lower.endswith(".xlsx") and "cv" not in lower:
                target = name
                break
    if not target:
        print("ERROR: Could not find Table 14 annual pay file.")
        print("Files in ZIP:")
        for n in sorted(zf.namelist()):
            if n.lower().endswith(".xlsx"):
                print(f"  {n}")
        return {}

    print(f"Table 14 target: {os.path.basename(target)}")

    raw_path = SCRIPT_DIR / os.path.basename(target)
    with open(raw_path, "wb") as f:
        f.write(zf.read(target))

    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)

    # Table 14 should have an "All" sheet too
    sheet_name = "All"
    if sheet_name not in wb.sheetnames:
        # Fallback: first non-Notes sheet
        for name in wb.sheetnames:
            if name.lower() != "notes":
                sheet_name = name
                break

    print(f"  Using sheet: '{sheet_name}'")
    ws = wb[sheet_name]

    header_row = detect_header(ws)
    if header_row < 0:
        print("  ERROR: Header row not found in Table 14.")
        wb.close()
        return {}

    cols = find_percentile_cols(ws, header_row)
    print(f"  Header row {header_row}, columns: {cols}")

    uk = {}
    c_desc = cols["desc"]
    c_code = cols["code"]
    c_median = cols["median"]
    c_p25 = cols.get("p25")
    c_p75 = cols.get("p75")

    for row in ws.iter_rows(min_row=header_row + 1, max_col=25, values_only=True):
        code_val = row[c_code] if c_code is not None and c_code < len(row) else None
        soc = is_4digit_soc(code_val)
        if not soc:
            continue

        desc_raw = row[c_desc] if c_desc is not None and c_desc < len(row) else None
        occ_title = str(desc_raw).strip() if desc_raw else ""

        # Table 14 is UK-level — no region prefix to strip, but the description
        # may still have leading spaces
        occ_title = occ_title.lstrip()

        p25 = parse_salary(row[c_p25]) if c_p25 is not None and c_p25 < len(row) else None
        p50 = parse_salary(row[c_median]) if c_median < len(row) else None
        p75 = parse_salary(row[c_p75]) if c_p75 is not None and c_p75 < len(row) else None

        uk[soc] = {"description": occ_title, "p25": p25, "p50": p50, "p75": p75}

    wb.close()
    return uk


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_json(data: dict, filename: str) -> None:
    filepath = SCRIPT_DIR / filename
    sorted_data = dict(sorted(data.items()))
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(sorted_data, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {len(sorted_data)} SOC codes -> {filepath.name}")


def main():
    # --- Table 15 → Scotland ---
    print("=" * 60)
    print("ASHE Table 15 (region x 4-digit SOC) -> Scotland")
    print("=" * 60)
    try:
        t15_bytes = download_zip(TABLE15_URL)
        t15_zip = zipfile.ZipFile(t15_bytes)
        scotland = parse_table15_scotland(t15_zip)
    except Exception as e:
        print(f"ERROR: {e}")
        scotland = {}

    # --- Table 14 → UK ---
    print()
    print("=" * 60)
    print("ASHE Table 14 (UK-level 4-digit SOC) -> UK fallback")
    print("=" * 60)
    try:
        t14_bytes = download_zip(TABLE14_URL)
        t14_zip = zipfile.ZipFile(t14_bytes)
        uk = parse_table14_uk(t14_zip)
    except Exception as e:
        print(f"ERROR: {e}")
        uk = {}

    # --- Results ---
    print()
    print("=" * 60)
    print("Results")
    print("=" * 60)

    scot_med = sum(1 for v in scotland.values() if v["p50"] is not None)
    scot_p25 = sum(1 for v in scotland.values() if v["p25"] is not None)
    uk_med = sum(1 for v in uk.values() if v["p50"] is not None)
    uk_p25 = sum(1 for v in uk.values() if v["p25"] is not None)

    print(f"  Scotland: {len(scotland)} SOC codes ({scot_med} with median, {scot_p25} with p25)")
    print(f"  UK:       {len(uk)} SOC codes ({uk_med} with median, {uk_p25} with p25)")

    if scotland:
        sample = list(scotland.items())[:3]
        print("\n  Scotland sample:")
        for soc, e in sample:
            print(f"    {soc}: {e}")

    if uk:
        sample = list(uk.items())[:3]
        print("\n  UK sample:")
        for soc, e in sample:
            print(f"    {soc}: {e}")

    # --- Write JSON ---
    print()
    if scotland:
        write_json(scotland, "ashe-2025-scotland.json")
    else:
        print("  WARNING: No Scotland data!")

    if uk:
        write_json(uk, "ashe-2025-uk.json")
    else:
        print("  WARNING: No UK data!")

    if not scotland and not uk:
        print("\nERROR: No data extracted.")
        sys.exit(1)

    print(f"\nSource: ASHE 2025 provisional (ONS, released Oct 2025)")
    print("Done.")


if __name__ == "__main__":
    main()
