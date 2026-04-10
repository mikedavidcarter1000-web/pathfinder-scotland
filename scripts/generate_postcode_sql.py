#!/usr/bin/env python3
"""
Generate SQL migration file for Scottish SIMD postcodes.
Reads from the official SIMD 2020v2 Excel file and generates SQL INSERT statements.
"""

import openpyxl
import os
import sys

def escape_sql_string(s):
    """Escape single quotes for SQL."""
    if s is None:
        return None
    return str(s).replace("'", "''")

def normalize_postcode(postcode):
    """Normalize postcode by removing spaces and converting to uppercase."""
    if postcode is None:
        return None
    return postcode.upper().replace(' ', '')

def main():
    excel_file = os.path.join(os.path.dirname(__file__), '..', 'simd_postcodes.xlsx')
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'supabase', 'migrations')

    print(f"Loading Excel file: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['All postcodes']

    # Get total rows for progress tracking
    total_rows = ws.max_row - 1  # Minus header
    print(f"Total postcodes to process: {total_rows}")

    # Create migration file
    timestamp = "20240129000001"
    migration_file = os.path.join(output_dir, f"{timestamp}_simd_postcodes_full.sql")

    print(f"Writing to: {migration_file}")

    with open(migration_file, 'w', encoding='utf-8') as f:
        f.write("-- SIMD 2020v2 Full Postcode Lookup Data\n")
        f.write("-- Source: Scottish Government SIMD 2020v2 Postcode Lookup\n")
        f.write("-- https://www.gov.scot/publications/scottish-index-of-multiple-deprivation-2020v2-postcode-look-up/\n")
        f.write(f"-- Total postcodes: {total_rows}\n\n")

        # First, delete existing sample data
        f.write("-- Delete existing sample SIMD postcodes\n")
        f.write("DELETE FROM simd_postcodes;\n\n")

        # Process in batches for better performance
        batch_size = 1000
        batch = []
        processed = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i == 0:
                continue  # Skip if somehow got header

            postcode = row[0]
            datazone = row[1]
            simd_decile = row[4]  # SIMD2020_Decile column

            # Validate data
            if not postcode or simd_decile is None:
                skipped += 1
                continue

            try:
                decile = int(simd_decile)
                if decile < 1 or decile > 10:
                    skipped += 1
                    continue
            except (ValueError, TypeError):
                skipped += 1
                continue

            normalized = normalize_postcode(postcode)
            dz_escaped = escape_sql_string(datazone) if datazone else None

            if dz_escaped:
                batch.append(f"('{normalized}', {decile}, '{dz_escaped}')")
            else:
                batch.append(f"('{normalized}', {decile}, NULL)")

            processed += 1

            # Write batch
            if len(batch) >= batch_size:
                f.write(f"INSERT INTO simd_postcodes (postcode, simd_decile, datazone) VALUES\n")
                f.write(",\n".join(batch))
                f.write("\nON CONFLICT (postcode) DO UPDATE SET simd_decile = EXCLUDED.simd_decile, datazone = EXCLUDED.datazone;\n\n")
                batch = []

                # Progress update
                if processed % 10000 == 0:
                    print(f"Processed: {processed}/{total_rows} ({100*processed/total_rows:.1f}%)")

        # Write remaining batch
        if batch:
            f.write(f"INSERT INTO simd_postcodes (postcode, simd_decile, datazone) VALUES\n")
            f.write(",\n".join(batch))
            f.write("\nON CONFLICT (postcode) DO UPDATE SET simd_decile = EXCLUDED.simd_decile, datazone = EXCLUDED.datazone;\n\n")

        # Add indexes if not exist (should already exist from initial migration)
        f.write("-- Ensure indexes exist\n")
        f.write("CREATE INDEX IF NOT EXISTS idx_simd_postcodes_postcode ON simd_postcodes(postcode);\n")
        f.write("CREATE INDEX IF NOT EXISTS idx_simd_postcodes_decile ON simd_postcodes(simd_decile);\n")

    print(f"\nDone! Processed: {processed}, Skipped: {skipped}")
    print(f"Migration file: {migration_file}")

    wb.close()

if __name__ == "__main__":
    main()
